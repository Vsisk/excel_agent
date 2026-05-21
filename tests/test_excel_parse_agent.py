import re
from pathlib import Path

import pytest
from openpyxl import Workbook
import xlwt

from agent.excel_agent.excel_parse_handler import handle_excel_parse
from agent.excel_agent.excel_reader import ExcelReader
from agent.excel_agent.excel_visualizer import ExcelVisualizer
from agent.excel_agent.llm_sheet_grouper import LLMSheetGrouper
from agent.excel_agent.logic_builder import LogicBuilder
from agent.excel_agent.models import (
    ALLOWED_LOGIC_PAGE_NAMES,
    CellRange,
    ExcelRegion,
    RegionGroup,
    RegionSnapshot,
    SheetGrouping,
    gen_id,
)
from agent.excel_agent.grouping_memory import WorkbookGroupingMemory
from agent.excel_agent.region_classifier import RegionClassifier
from agent.excel_agent.region_markdown_builder import RegionMarkdownBuilder
from agent.excel_agent.region_splitter import RegionSplitter
from agent.excel_agent.sheet_profiler import SheetProfiler


def test_gen_id_uses_date_and_eight_digits():
    value = gen_id()

    assert re.fullmatch(r"\d{16}", value)


def test_region_model_uses_one_based_range():
    region = ExcelRegion(
        sheet_id="sheet_1",
        cell_range=CellRange(start_row=1, end_row=3, start_col=1, end_col=2),
        raw_text=["A B", "C D"],
    )

    assert region.cell_range.start_row == 1
    assert region.cell_range.start_col == 1


def test_grouping_models_capture_minimal_llm_contract():
    snapshot = RegionSnapshot(
        region_id="region_1",
        sheet_id="sheet_1",
        cell_range=CellRange(1, 2, 1, 2),
        markdown="| A | B |\n| --- | --- |\n| x | y |",
        raw_text=["A B", "x y"],
        rule_classification={"logic_area_type": "fields", "confidence": 0.82},
        truncated=False,
    )
    grouping = SheetGrouping(
        logic_page_name="bill_summary_page",
        groups=[RegionGroup(region_ids=["region_1"], reason="single section")],
    )

    assert "bill_charge_page" in ALLOWED_LOGIC_PAGE_NAMES
    assert snapshot.region_id == grouping.groups[0].region_ids[0]
    assert grouping.logic_page_name == "bill_summary_page"


def make_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Invoice"
    ws["B1"] = "Amount"
    ws["A2"] = "INV-001"
    ws["B2"] = 120
    detail = wb.create_sheet("Details")
    detail["A1"] = "Item"
    detail["B1"] = "Qty"
    detail["A2"] = "Storage"
    detail["B2"] = 3
    wb.save(path)


def make_xls_workbook(path: Path) -> None:
    wb = xlwt.Workbook()
    summary = wb.add_sheet("Summary")
    summary.write(0, 0, "Invoice")
    summary.write(0, 1, "Amount")
    summary.write(1, 0, "INV-001")
    summary.write(1, 1, 120)
    detail = wb.add_sheet("Details")
    detail.write(0, 0, "Item")
    detail.write(0, 1, "Qty")
    detail.write(1, 0, "Storage")
    detail.write(1, 1, 3)
    wb.save(str(path))


def test_excel_reader_reads_sheet_metadata_and_rows(tmp_path):
    path = tmp_path / "sample.xlsx"
    make_workbook(path)

    reader = ExcelReader(str(path))
    workbook = reader.read()

    assert [sheet["sheet_name"] for sheet in workbook["sheet_list"]] == ["Summary", "Details"]
    assert workbook["sheet_list"][0]["sheet_index"] == 0
    assert workbook["sheet_list"][0]["max_row"] == 2
    assert workbook["sheet_list"][0]["max_col"] == 2
    rows = list(reader.iter_rows("Summary", min_row=1, max_row=2, min_col=1, max_col=2))
    reader.close()

    assert rows == [["Invoice", "Amount"], ["INV-001", 120]]


def test_excel_reader_reads_legacy_xls_metadata_and_rows(tmp_path):
    path = tmp_path / "legacy.xls"
    make_xls_workbook(path)

    reader = ExcelReader(str(path))
    workbook = reader.read()

    assert [sheet["sheet_name"] for sheet in workbook["sheet_list"]] == ["Summary", "Details"]
    assert workbook["sheet_list"][0]["sheet_index"] == 0
    assert workbook["sheet_list"][0]["max_row"] == 2
    assert workbook["sheet_list"][0]["max_col"] == 2
    rows = list(reader.iter_rows("Summary", min_row=1, max_row=2, min_col=1, max_col=2))
    reader.close()

    assert rows == [["Invoice", "Amount"], ["INV-001", 120]]


def test_excel_reader_reads_only_requested_range(tmp_path):
    path = tmp_path / "range.xlsx"
    make_workbook(path)
    reader = ExcelReader(str(path))
    reader.read()

    rows = reader.read_range("Summary", CellRange(1, 2, 1, 2))
    reader.close()

    assert rows == [["Invoice", "Amount"], ["INV-001", 120]]


def test_region_markdown_builder_preserves_table_and_truncates():
    region = ExcelRegion("sheet_1", CellRange(1, 13, 1, 2), raw_text=[])
    rows = [["Item", "Amount"]] + [[f"row-{i}", i] for i in range(1, 13)]

    snapshot = RegionMarkdownBuilder.build_region_snapshot(
        region_id="region_1",
        region=region,
        rows=rows,
        max_instance_rows=10,
    )

    assert "| Item | Amount |" in snapshot.markdown
    assert "| --- | --- |" not in snapshot.markdown
    assert "| row-10 | 10 |" in snapshot.markdown
    assert "row-11" not in snapshot.markdown
    assert "truncated after 10 rows" in snapshot.markdown
    assert snapshot.truncated is True
    assert snapshot.rule_classification == {"logic_area_type": "unknown", "confidence": 1.0}


def test_excel_pipeline_preserves_low_resource_language_text(tmp_path):
    path = tmp_path / "unicode_languages.xlsx"
    thai_customer = "\u0e25\u0e39\u0e01\u0e04\u0e49\u0e32"
    thai_test = "\u0e17\u0e14\u0e2a\u0e2d\u0e1a"
    arabic_amount = "\u0627\u0644\u0645\u0628\u0644\u063a"
    arabic_hello = "\u0645\u0631\u062d\u0628\u0627"

    wb = Workbook()
    ws = wb.active
    ws.title = "Unicode"
    ws["A1"] = thai_customer
    ws["B1"] = arabic_amount
    ws["A2"] = thai_test
    ws["B2"] = arabic_hello
    wb.save(path)

    reader = ExcelReader(str(path))
    sheet_info = reader.read()["sheet_list"][0]
    rows = reader.read_range(sheet_info["sheet_name"], CellRange(1, 2, 1, 2))
    profile = SheetProfiler.profile(reader, sheet_info)
    regions = RegionSplitter.split(reader, sheet_info, profile)
    snapshot = RegionMarkdownBuilder.build_region_snapshot(
        region_id="region_1",
        region=regions[0],
        rows=reader.read_range(sheet_info["sheet_name"], regions[0].cell_range),
    )
    reader.close()

    assert rows == [[thai_customer, arabic_amount], [thai_test, arabic_hello]]
    assert regions[0].raw_text == [
        f"{thai_customer} {arabic_amount}",
        f"{thai_test} {arabic_hello}",
    ]
    assert snapshot.markdown == (
        f"| {thai_customer} | {arabic_amount} |\n"
        f"| {thai_test} | {arabic_hello} |"
    )


def test_llm_sheet_grouper_validates_groups_and_page_name():
    calls = []

    def fake_generate(prompt_template, llm_name="base", lang="zh", **kwargs):
        calls.append((prompt_template, llm_name, kwargs))
        return {
            "logic_page_name": "bill_charge_page",
            "groups": [
                {"region_ids": ["region_1", "missing", "region_1"], "reason": "charges"},
            ],
        }

    snapshots = [
        RegionSnapshot(
            "region_1",
            "sheet_1",
            CellRange(1, 2, 1, 2),
            "r1",
            [],
            {"logic_area_type": "fields", "confidence": 0.82},
            False,
        ),
        RegionSnapshot(
            "region_2",
            "sheet_1",
            CellRange(5, 6, 1, 3),
            "r2",
            [],
            {"logic_area_type": "fee_table", "confidence": 0.78},
            False,
        ),
    ]
    grouping = LLMSheetGrouper(llm_generate=fake_generate).group_sheet(
        sheet_info={
            "sheet_id": "sheet_1",
            "sheet_name": "Charges",
            "sheet_index": 0,
            "max_row": 6,
            "max_col": 3,
            "merged_cells": [],
        },
        region_snapshots=snapshots,
        visual_summaries=[],
    )

    assert grouping.logic_page_name == "bill_charge_page"
    assert [group.region_ids for group in grouping.groups] == [["region_1"], ["region_2"]]
    assert calls[0][0] == "excel_sheet_grouping"
    assert calls[0][1] == "base"


def test_llm_sheet_grouper_falls_back_on_failure():
    def failing_generate(*args, **kwargs):
        raise RuntimeError("boom")

    snapshots = [
        RegionSnapshot(
            "region_1",
            "sheet_1",
            CellRange(1, 1, 1, 1),
            "r1",
            [],
            {"logic_area_type": "unknown", "confidence": 0.35},
            False,
        )
    ]
    grouper = LLMSheetGrouper(llm_generate=failing_generate)

    grouping = grouper.group_sheet(
        sheet_info={
            "sheet_id": "sheet_1",
            "sheet_name": "Unknown",
            "sheet_index": 0,
            "max_row": 1,
            "max_col": 1,
            "merged_cells": [],
        },
        region_snapshots=snapshots,
        visual_summaries=[],
    )

    assert grouping.logic_page_name == "bill_summary_page"
    assert grouping.groups[0].region_ids == ["region_1"]
    assert "fallback" in grouping.groups[0].reason
    assert grouper.last_fallback_reason == "boom"


def test_workbook_grouping_memory_retrieves_global_group_templates():
    vectors = {
        "customer amount": [1.0, 0.0],
        "storage usage": [0.0, 1.0],
        "current customer amount": [1.0, 0.0],
    }

    def fake_embedding(text: str) -> list[float]:
        for key, vector in vectors.items():
            if key in text:
                return vector
        return [0.2, 0.8]

    memory = WorkbookGroupingMemory(embedding_generate=fake_embedding, top_k=1)
    snapshot = RegionSnapshot(
        "region_1",
        "sheet_1",
        CellRange(1, 2, 1, 2),
        "customer amount",
        ["customer amount"],
        {"logic_area_type": "unknown", "confidence": 1.0},
        False,
    )
    memory.remember_sheet_grouping(
        sheet_info={
            "sheet_id": "sheet_1",
            "sheet_name": "Summary",
            "sheet_index": 0,
            "max_row": 2,
            "max_col": 2,
            "merged_cells": [],
        },
        region_snapshots=[snapshot],
        visual_summaries=[],
        grouping=SheetGrouping(
            logic_page_name="bill_summary_page",
            groups=[RegionGroup(region_ids=["region_1"], reason="customer amount group")],
        ),
    )

    matches = memory.retrieve_for_sheet(
        sheet_info={
            "sheet_id": "sheet_2",
            "sheet_name": "Next",
            "sheet_index": 1,
            "max_row": 2,
            "max_col": 2,
            "merged_cells": [],
        },
        region_snapshots=[
            RegionSnapshot(
                "region_1",
                "sheet_2",
                CellRange(1, 2, 1, 2),
                "current customer amount",
                ["current customer amount"],
                {"logic_area_type": "unknown", "confidence": 1.0},
                False,
            )
        ],
        visual_summaries=[],
    )

    assert matches == [
        {
            "template_id": "template_1",
            "similarity": 1.0,
            "support_count": 1,
            "sample_sheet_names": ["Summary"],
            "template_preview": "customer amount group | regions=1 | customer amount",
        }
    ]


def test_visualizer_calls_vl_for_low_confidence_region():
    calls = []

    def fake_generate(prompt_template, llm_name="base", lang="zh", **kwargs):
        calls.append((prompt_template, llm_name, kwargs))
        return {"target_id": kwargs["target_id"], "summary": "looks like notes", "confidence": 0.8}

    snapshot = RegionSnapshot(
        "region_1",
        "sheet_1",
        CellRange(1, 1, 1, 1),
        "unclear",
        ["unclear"],
        {"logic_area_type": "unknown", "confidence": 0.35},
        False,
    )

    visualizer = ExcelVisualizer(llm_generate=fake_generate)
    result = visualizer.collect_visual_summaries(
        file_uri="unused.xlsx",
        sheet_info={
            "sheet_id": "sheet_1",
            "sheet_name": "S",
            "sheet_index": 0,
            "max_row": 1,
            "max_col": 1,
            "merged_cells": [],
        },
        region_snapshots=[snapshot],
    )

    assert result["summaries"][0]["target_id"] == "region_1"
    assert calls[0][0] == "excel_visual_summary"
    assert calls[0][1] == "vl"


def test_visualizer_skips_high_confidence_region():
    def fail_generate(*args, **kwargs):
        raise AssertionError("vl should not be called")

    snapshot = RegionSnapshot(
        "region_1",
        "sheet_1",
        CellRange(1, 2, 1, 2),
        "clear",
        ["clear"],
        {"logic_area_type": "fields", "confidence": 0.82},
        False,
    )

    result = ExcelVisualizer(llm_generate=fail_generate).collect_visual_summaries(
        file_uri="unused.xlsx",
        sheet_info={
            "sheet_id": "sheet_1",
            "sheet_name": "S",
            "sheet_index": 0,
            "max_row": 2,
            "max_col": 2,
            "merged_cells": [],
        },
        region_snapshots=[snapshot],
    )

    assert result == {"summaries": [], "skipped": []}


def test_sheet_profiler_computes_used_range_and_density(tmp_path):
    path = tmp_path / "profile.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Profile"
    ws["B2"] = "Name"
    ws["C2"] = "Amount"
    ws["B3"] = "A"
    ws["C3"] = 10
    ws["B6"] = "Footer"
    wb.save(path)

    reader = ExcelReader(str(path))
    workbook = reader.read()
    sheet_info = workbook["sheet_list"][0]

    profile = SheetProfiler.profile(reader, sheet_info)
    reader.close()

    assert profile.used_range.to_dict() == {
        "start_row": 2,
        "end_row": 6,
        "start_col": 2,
        "end_col": 3,
    }
    assert profile.row_density == [0, 2, 2, 0, 0, 1]
    assert profile.col_density == [0, 3, 2]


def test_region_splitter_splits_by_empty_rows_and_columns(tmp_path):
    path = tmp_path / "regions.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Regions"
    ws["A1"] = "Customer"
    ws["B1"] = "ACME"
    ws["A2"] = "Billing Month"
    ws["B2"] = "2026-04"
    ws["A5"] = "Item"
    ws["B5"] = "Qty"
    ws["C5"] = "Amount"
    ws["A6"] = "Compute"
    ws["B6"] = 2
    ws["C6"] = 100
    ws["F5"] = "Notes"
    ws["F6"] = "Long running text"
    wb.save(path)

    reader = ExcelReader(str(path))
    sheet_info = reader.read()["sheet_list"][0]
    profile = SheetProfiler.profile(reader, sheet_info)

    regions = RegionSplitter.split(reader, sheet_info, profile)
    reader.close()

    assert [region.cell_range.to_dict() for region in regions] == [
        {"start_row": 1, "end_row": 2, "start_col": 1, "end_col": 2},
        {"start_row": 5, "end_row": 6, "start_col": 1, "end_col": 3},
        {"start_row": 5, "end_row": 6, "start_col": 6, "end_col": 6},
    ]
    assert regions[1].raw_text == ["Item Qty Amount", "Compute 2 100"]


def test_region_splitter_splits_by_two_consecutive_empty_rows(tmp_path):
    path = tmp_path / "empty_rows.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Rows"
    ws["A1"] = "Customer"
    ws["B1"] = "ACME"
    ws["A2"] = "Month"
    ws["B2"] = "2026-05"
    ws["A5"] = "Item"
    ws["B5"] = "Amount"
    ws["A6"] = "Compute"
    ws["B6"] = 100
    wb.save(path)

    reader = ExcelReader(str(path))
    sheet_info = reader.read()["sheet_list"][0]
    profile = SheetProfiler.profile(reader, sheet_info)

    regions = RegionSplitter.split(reader, sheet_info, profile)
    reader.close()

    assert [region.cell_range.to_dict() for region in regions] == [
        {"start_row": 1, "end_row": 2, "start_col": 1, "end_col": 2},
        {"start_row": 5, "end_row": 6, "start_col": 1, "end_col": 2},
    ]


def test_region_splitter_splits_by_two_consecutive_empty_columns(tmp_path):
    path = tmp_path / "empty_cols.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Columns"
    ws["A1"] = "Customer"
    ws["B1"] = "ACME"
    ws["A2"] = "Month"
    ws["B2"] = "2026-05"
    ws["E1"] = "Item"
    ws["F1"] = "Amount"
    ws["E2"] = "Compute"
    ws["F2"] = 100
    wb.save(path)

    reader = ExcelReader(str(path))
    sheet_info = reader.read()["sheet_list"][0]
    profile = SheetProfiler.profile(reader, sheet_info)

    regions = RegionSplitter.split(reader, sheet_info, profile)
    reader.close()

    assert [region.cell_range.to_dict() for region in regions] == [
        {"start_row": 1, "end_row": 2, "start_col": 1, "end_col": 2},
        {"start_row": 1, "end_row": 2, "start_col": 5, "end_col": 6},
    ]


def test_region_splitter_splits_row_local_two_empty_column_gap(tmp_path):
    path = tmp_path / "row_local_empty_cols.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "RowGap"
    ws["A1"] = "Customer"
    ws["B1"] = "ACME"
    ws["E1"] = "Item"
    ws["F1"] = "Amount"
    ws["A2"] = "Month"
    ws["B2"] = "2026-05"
    ws["E2"] = "Compute"
    ws["F2"] = 100
    ws["C3"] = "Tax"
    ws["D3"] = 8
    wb.save(path)

    reader = ExcelReader(str(path))
    sheet_info = reader.read()["sheet_list"][0]
    profile = SheetProfiler.profile(reader, sheet_info)

    regions = RegionSplitter.split(reader, sheet_info, profile)
    reader.close()

    assert [region.cell_range.to_dict() for region in regions] == [
        {"start_row": 1, "end_row": 2, "start_col": 1, "end_col": 2},
        {"start_row": 1, "end_row": 2, "start_col": 5, "end_col": 6},
        {"start_row": 3, "end_row": 3, "start_col": 3, "end_col": 4},
    ]


def test_region_splitter_keeps_trailing_merged_summary_row_together(tmp_path):
    path = tmp_path / "merged_summary.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Item"
    ws["B1"] = "Qty"
    ws["C1"] = "Amount"
    ws["A2"] = "Compute"
    ws["B2"] = 2
    ws["C2"] = 100
    ws["A3"] = "Storage"
    ws["B3"] = 3
    ws["C3"] = 240
    ws.merge_cells("A4:C4")
    ws["A4"] = "Grand total"
    ws["F4"] = 5
    ws["G4"] = 340
    wb.save(path)

    reader = ExcelReader(str(path))
    sheet_info = reader.read()["sheet_list"][0]
    profile = SheetProfiler.profile(reader, sheet_info)

    regions = RegionSplitter.split(reader, sheet_info, profile)
    reader.close()

    assert [region.cell_range.to_dict() for region in regions] == [
        {"start_row": 1, "end_row": 3, "start_col": 1, "end_col": 3},
        {"start_row": 4, "end_row": 4, "start_col": 1, "end_col": 7},
    ]
    assert regions[1].raw_text == ["Grand total 5 340"]


def test_region_splitter_splits_on_row_density_jump(tmp_path):
    path = tmp_path / "density_jump.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Density"
    ws["A1"] = "Invoice"
    ws["A2"] = "Item"
    ws["B2"] = "Qty"
    ws["C2"] = "Amount"
    ws["D2"] = "Tax"
    ws["E2"] = "Total"
    ws["A3"] = "Compute"
    ws["B3"] = 2
    ws["C3"] = 100
    ws["D3"] = 8
    ws["E3"] = 108
    wb.save(path)

    reader = ExcelReader(str(path))
    sheet_info = reader.read()["sheet_list"][0]
    profile = SheetProfiler.profile(reader, sheet_info)

    regions = RegionSplitter.split(reader, sheet_info, profile)
    reader.close()

    assert [region.cell_range.to_dict() for region in regions] == [
        {"start_row": 1, "end_row": 1, "start_col": 1, "end_col": 1},
        {"start_row": 2, "end_row": 3, "start_col": 1, "end_col": 5},
    ]


def test_region_splitter_returns_no_regions_for_empty_sheet(tmp_path):
    path = tmp_path / "empty_split.xlsx"
    wb = Workbook()
    wb.active.title = "Empty"
    wb.save(path)

    reader = ExcelReader(str(path))
    sheet_info = reader.read()["sheet_list"][0]
    profile = SheetProfiler.profile(reader, sheet_info)

    regions = RegionSplitter.split(reader, sheet_info, profile)
    reader.close()

    assert regions == []


def test_region_classifier_is_neutral_legacy_adapter():
    region = ExcelRegion(
        sheet_id="sheet_1",
        cell_range=CellRange(1, 2, 1, 2),
        raw_text=["Customer ACME", "Billing Month 2026-04"],
    )

    assert RegionClassifier.classify(region) == {
        "logic_area_type": "unknown",
        "confidence": 1.0,
    }
    assert not hasattr(RegionClassifier, "_looks_like_fields")
    assert not hasattr(RegionClassifier, "_has_clear_header")
    assert not hasattr(RegionClassifier, "_looks_like_fee_table")
    assert not hasattr(RegionClassifier, "_looks_like_plain_text")


def test_handle_excel_parse_does_not_call_rule_classifier(tmp_path, monkeypatch):
    path = tmp_path / "no_classifier.xlsx"
    make_workbook(path)

    def fail_classify(*args, **kwargs):
        raise AssertionError("RegionClassifier.classify should not be called")

    def fake_generate(prompt_template, llm_name="base", lang="zh", **kwargs):
        if prompt_template == "excel_sheet_grouping":
            return {
                "logic_page_name": "bill_summary_page",
                "groups": [{"region_ids": ["region_1"], "reason": "single structural region"}],
            }
        return {"target_id": kwargs["target_id"], "summary": "", "confidence": 0.0}

    monkeypatch.setattr(RegionClassifier, "classify", fail_classify)

    response = handle_excel_parse(
        {
            "request_type": "EXCEL_PARSE",
            "task_id": "task_no_classifier",
            "site_id": "site_1",
            "project_id": "project_1",
            "payload": {"excel_instance_id": "excel_1", "file_uri": str(path), "parse_mode": "full"},
        },
        llm_generate=fake_generate,
    )

    assert response["status"] == "success"


def test_logic_builder_creates_page_and_area():
    sheet_info = {
        "sheet_id": "sheet_1",
        "sheet_name": "Summary",
        "sheet_index": 0,
        "max_row": 2,
        "max_col": 2,
        "merged_cells": [],
    }
    region = ExcelRegion(
        sheet_id="sheet_1",
        cell_range=CellRange(1, 2, 1, 2),
        raw_text=["Customer ACME", "Billing Month 2026-04"],
    )

    page = LogicBuilder.build_page("excel_1", sheet_info)
    area = LogicBuilder.build_area(
        "excel_1",
        sheet_info,
        region,
        {"logic_area_type": "fields", "confidence": 0.82},
    )

    assert page["logic_page_relation"] == {
        "type": "excel",
        "excel_instance_id": "excel_1",
        "sheet_id": "sheet_1",
        "sheet_name": "Summary",
        "sheet_index": 0,
    }
    assert area["logic_area_type"] == "fields"
    assert area["location_list"][0]["cell_range"] == {
        "start_row": 1,
        "end_row": 2,
        "start_col": 1,
        "end_col": 2,
    }
    assert area["location_list"][0]["raw_excel_text_list"] == region.raw_text


def test_logic_builder_creates_named_page_and_grouped_area():
    sheet_info = {
        "sheet_id": "sheet_1",
        "sheet_name": "Charges",
        "sheet_index": 0,
        "max_row": 6,
        "max_col": 3,
        "merged_cells": [],
    }
    region1 = ExcelRegion("sheet_1", CellRange(1, 2, 1, 2), ["A B"])
    region2 = ExcelRegion("sheet_1", CellRange(5, 6, 1, 3), ["Item Qty Amount"])
    classifications = {
        "region_1": {"logic_area_type": "fields", "confidence": 0.82},
        "region_2": {"logic_area_type": "fee_table", "confidence": 0.78},
    }

    page = LogicBuilder.build_page("excel_1", sheet_info, logic_page_name="bill_charge_page")
    area = LogicBuilder.build_grouped_area(
        "excel_1",
        sheet_info,
        group=RegionGroup(region_ids=["region_1", "region_2"], reason="same page content"),
        region_by_id={"region_1": region1, "region_2": region2},
        classification_by_id=classifications,
    )

    assert page["logic_page_name"] == "bill_charge_page"
    assert area["logic_area_type"] == "fee_table"
    assert len(area["location_list"]) == 2
    assert area["group_reason"] == "same page content"


def test_handle_excel_parse_returns_ws_result_for_multiple_sheets(tmp_path):
    path = tmp_path / "full.xlsx"
    make_workbook(path)

    response = handle_excel_parse(
        {
            "request_type": "EXCEL_PARSE",
            "task_id": "task_1",
            "site_id": "site_1",
            "project_id": "project_1",
            "payload": {
                "excel_instance_id": "excel_1",
                "file_uri": str(path),
                "parse_mode": "full",
            },
        }
    )

    assert response["request_type"] == "EXCEL_PARSE_RESULT"
    assert response["task_id"] == "task_1"
    assert response["status"] == "success"
    assert len(response["payload"]["logic_page_list"]) == 2
    assert len(response["payload"]["logic_area_list"]) >= 2
    assert response["payload"]["parse_index"]["excel_instance_id"] == "excel_1"
    assert response["payload"]["parse_index"]["sheet_count"] == 2
    assert response["payload"]["parse_index"]["area_count"] == len(response["payload"]["logic_area_list"])


def test_handle_excel_parse_uses_sheet_grouping_llm(tmp_path):
    path = tmp_path / "grouped.xlsx"
    make_workbook(path)

    def fake_generate(prompt_template, llm_name="base", lang="zh", **kwargs):
        if prompt_template == "excel_sheet_grouping":
            return {
                "logic_page_name": "bill_charge_page",
                "groups": [{"region_ids": ["region_1"], "reason": "summary fields"}],
            }
        return {"target_id": kwargs["target_id"], "summary": "", "confidence": 0.0}

    response = handle_excel_parse(
        {
            "request_type": "EXCEL_PARSE",
            "task_id": "task_grouped",
            "site_id": "site_1",
            "project_id": "project_1",
            "payload": {"excel_instance_id": "excel_1", "file_uri": str(path), "parse_mode": "full"},
        },
        llm_generate=fake_generate,
    )

    assert response["status"] == "success"
    assert response["payload"]["logic_page_list"][0]["logic_page_name"] == "bill_charge_page"
    assert response["payload"]["parse_index"]["llm_used"] is True
    assert response["payload"]["parse_index"]["sheet_grouping_count"] == 2


def test_handle_excel_parse_outputs_sheet_content_blocks(tmp_path):
    path = tmp_path / "sheet_content.xlsx"
    make_workbook(path)

    def fake_generate(prompt_template, llm_name="base", lang="zh", **kwargs):
        if prompt_template == "excel_sheet_grouping":
            return {
                "logic_page_name": "bill_charge_page",
                "groups": [{"region_ids": ["region_1"], "reason": "sheet block"}],
            }
        return {"target_id": kwargs["target_id"], "summary": "", "confidence": 0.0}

    response = handle_excel_parse(
        {
            "request_type": "EXCEL_PARSE",
            "task_id": "task_sheet_content",
            "site_id": "site_1",
            "project_id": "project_1",
            "payload": {"excel_instance_id": "excel_1", "file_uri": str(path), "parse_mode": "full"},
        },
        llm_generate=fake_generate,
    )

    sheet_content = response["payload"]["sheet_content"]

    assert sheet_content[0]["page_id"] == 1
    assert sheet_content[0]["page_type"] == "bill_charge_page"
    assert sheet_content[0]["blocks"][0]["group_id"] == "group_1"
    assert sheet_content[0]["blocks"][0]["bbox"] == {
        "left": 1,
        "right": 2,
        "top": 1,
        "bottom": 2,
    }
    assert "| Invoice | Amount |" in sheet_content[0]["blocks"][0]["table_md"]
    assert sheet_content[1]["page_id"] == 2


def test_handle_excel_parse_combines_group_table_md_as_pipe_rows(tmp_path):
    path = tmp_path / "combined_table_md.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined"
    ws["A1"] = "Invoice"
    ws["B1"] = "Amount"
    ws["A2"] = "INV-001"
    ws["B2"] = 120
    ws["A5"] = "Item"
    ws["B5"] = "Qty"
    ws["A6"] = "Storage"
    ws["B6"] = 3
    wb.save(path)

    def fake_generate(prompt_template, llm_name="base", lang="zh", **kwargs):
        if prompt_template == "excel_sheet_grouping":
            return {
                "logic_page_name": "bill_charge_page",
                "groups": [{"region_ids": ["region_1", "region_2"], "reason": "same block"}],
            }
        return {"target_id": kwargs["target_id"], "summary": "", "confidence": 0.0}

    response = handle_excel_parse(
        {
            "request_type": "EXCEL_PARSE",
            "task_id": "task_combined_table_md",
            "site_id": "site_1",
            "project_id": "project_1",
            "payload": {"excel_instance_id": "excel_1", "file_uri": str(path), "parse_mode": "full"},
        },
        llm_generate=fake_generate,
    )

    table_md = response["payload"]["sheet_content"][0]["blocks"][0]["table_md"]

    assert table_md == "\n".join(
        [
            "| Invoice | Amount |",
            "| INV-001 | 120 |",
            "| Item | Qty |",
            "| Storage | 3 |",
        ]
    )
    assert "\n\n" not in table_md


def test_handle_excel_parse_feeds_global_grouping_memory_to_next_sheet(tmp_path):
    path = tmp_path / "memory.xlsx"
    make_workbook(path)
    grouping_payloads = []

    def fake_generate(prompt_template, llm_name="base", lang="zh", **kwargs):
        if prompt_template == "excel_sheet_grouping":
            grouping_payloads.append(kwargs["sheet_payload"])
            return {
                "logic_page_name": "bill_summary_page",
                "groups": [{"region_ids": ["region_1"], "reason": "invoice amount block"}],
            }
        return {"target_id": kwargs["target_id"], "summary": "", "confidence": 0.0}

    def fake_embedding(text: str) -> list[float]:
        if "Invoice" in text or "Amount" in text:
            return [1.0, 0.0]
        return [0.9, 0.1]

    response = handle_excel_parse(
        {
            "request_type": "EXCEL_PARSE",
            "task_id": "task_memory",
            "site_id": "site_1",
            "project_id": "project_1",
            "payload": {"excel_instance_id": "excel_1", "file_uri": str(path), "parse_mode": "full"},
        },
        llm_generate=fake_generate,
        embedding_generate=fake_embedding,
    )

    assert response["status"] == "success"
    assert '"grouping_memory_matches": []' in grouping_payloads[0]
    assert '"template_id": "template_1"' in grouping_payloads[1]
    parse_index = response["payload"]["parse_index"]
    assert parse_index["grouping_memory_enabled"] is True
    assert parse_index["grouping_memory_used"] is True
    assert parse_index["grouping_memory_template_count"] == 1
    assert parse_index["grouping_memory_matches"][1]["matches"][0]["template_id"] == "template_1"
    assert parse_index["memory_consistency_warnings"] == []


def test_handle_excel_parse_falls_back_when_llm_fails(tmp_path):
    path = tmp_path / "fallback.xlsx"
    make_workbook(path)

    def failing_generate(*args, **kwargs):
        raise RuntimeError("llm offline")

    response = handle_excel_parse(
        {
            "request_type": "EXCEL_PARSE",
            "task_id": "task_fallback",
            "site_id": "site_1",
            "project_id": "project_1",
            "payload": {"excel_instance_id": "excel_1", "file_uri": str(path), "parse_mode": "full"},
        },
        llm_generate=failing_generate,
    )

    assert response["status"] == "success"
    assert response["payload"]["logic_page_list"][0]["logic_page_name"] == "bill_summary_page"
    assert response["payload"]["parse_index"]["llm_used"] is False
    assert "llm offline" in response["payload"]["parse_index"]["llm_fallback_reason"]


def test_handle_excel_parse_rejects_wrong_request_type(tmp_path):
    path = tmp_path / "bad.xlsx"
    make_workbook(path)

    with pytest.raises(ValueError, match="request_type must be EXCEL_PARSE"):
        handle_excel_parse(
            {
                "request_type": "OTHER",
                "task_id": "task_1",
                "site_id": "site_1",
                "project_id": "project_1",
                "payload": {
                    "excel_instance_id": "excel_1",
                    "file_uri": str(path),
                    "parse_mode": "full",
                },
            }
        )


def test_empty_sheet_returns_page_without_area(tmp_path):
    path = tmp_path / "empty.xlsx"
    wb = Workbook()
    wb.active.title = "Empty"
    wb.save(path)

    response = handle_excel_parse(
        {
            "request_type": "EXCEL_PARSE",
            "task_id": "task_empty",
            "site_id": "site_1",
            "project_id": "project_1",
            "payload": {
                "excel_instance_id": "excel_empty",
                "file_uri": str(path),
                "parse_mode": "full",
            },
        }
    )

    assert response["status"] == "success"
    assert len(response["payload"]["logic_page_list"]) == 1
    assert response["payload"]["logic_area_list"] == []
