import json
from pathlib import Path

import numpy

for name, fallback in {
    "short": int,
    "ushort": int,
    "intc": int,
    "uintc": int,
    "int_": int,
    "uint": int,
    "longlong": int,
    "ulonglong": int,
    "half": float,
    "float16": float,
    "single": float,
    "double": float,
    "longdouble": float,
    "int8": int,
    "int16": int,
    "int32": int,
    "int64": int,
    "uint8": int,
    "uint16": int,
    "uint32": int,
    "uint64": int,
    "intp": int,
    "uintp": int,
    "float32": float,
    "float64": float,
    "bool_": bool,
    "floating": float,
    "integer": int,
}.items():
    if not hasattr(numpy, name):
        setattr(numpy, name, fallback)

from openpyxl import Workbook

from agent.excel_agent.excel_parse_handler import handle_excel_parse


def _make_main_path_workbook(path: Path) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Invoice"
    summary["B1"] = "Amount"
    summary["A2"] = "INV-001"
    summary["B2"] = 120

    charges = wb.create_sheet("Charges")
    charges["A1"] = "Item"
    charges["B1"] = "Qty"
    charges["C1"] = "Total"
    charges["A2"] = "Storage"
    charges["B2"] = 3
    charges["C2"] = 360

    wb.save(path)


def test_handle_excel_parse_main_path_returns_sheet_content_for_each_sheet(tmp_path):
    path = tmp_path / "main_path.xlsx"
    _make_main_path_workbook(path)
    llm_calls = []

    def fake_generate(prompt_template, llm_name="base", lang="zh", **kwargs):
        llm_calls.append((prompt_template, llm_name, lang, json.loads(kwargs["sheet_payload"])))
        return {
            "logic_page_name": "bill_charge_page",
            "groups": [{"region_ids": ["region_1"], "reason": "single table"}],
        }

    sheet_content = handle_excel_parse(
        {
            "request_type": "EXCEL_PARSE",
            "task_id": "task_main",
            "site_id": "site_1",
            "project_id": "project_1",
            "payload": {
                "excel_instance_id": "excel_1",
                "file_uri": str(path),
                "parse_mode": "full",
            },
        },
        llm_generate=fake_generate,
        embedding_generate=lambda text: [1.0, 0.0],
    )

    assert [page["page_id"] for page in sheet_content] == [1, 2]
    assert [page["page_type"] for page in sheet_content] == [
        "bill_charge_page",
        "bill_charge_page",
    ]
    assert sheet_content[0]["blocks"] == [
        {
            "group_id": "group_1",
            "bbox": {"left": 1, "right": 2, "top": 1, "bottom": 2},
            "table_md": "| Invoice | Amount |\n| INV-001 | 120 |",
        }
    ]
    assert sheet_content[1]["blocks"] == [
        {
            "group_id": "group_1",
            "bbox": {"left": 1, "right": 3, "top": 1, "bottom": 2},
            "table_md": "| Item | Qty | Total |\n| Storage | 3 | 360 |",
        }
    ]

    assert [call[0] for call in llm_calls] == ["excel_sheet_grouping", "excel_sheet_grouping"]
    assert all(call[1] == "base" and call[2] == "zh" for call in llm_calls)
    assert llm_calls[0][3]["sheet"]["sheet_name"] == "Summary"
    assert llm_calls[0][3]["regions"][0]["table_md"] == "| Invoice | Amount |\n| INV-001 | 120 |"
    assert llm_calls[1][3]["sheet"]["sheet_name"] == "Charges"


def test_handle_excel_parse_main_path_keeps_empty_sheet_as_empty_page(tmp_path):
    path = tmp_path / "empty_page.xlsx"
    wb = Workbook()
    wb.active.title = "Empty"
    wb.save(path)

    sheet_content = handle_excel_parse(
        {
            "request_type": "EXCEL_PARSE",
            "task_id": "task_empty",
            "site_id": "site_1",
            "project_id": "project_1",
            "payload": {
                "excel_instance_id": "excel_1",
                "file_uri": str(path),
                "parse_mode": "full",
            },
        },
        llm_generate=lambda *args, **kwargs: {
            "logic_page_name": "bill_summary_page",
            "groups": [],
        },
        embedding_generate=lambda text: [1.0, 0.0],
    )

    assert sheet_content == [{"page_id": 1, "page_type": "bill_summary_page", "blocks": []}]
