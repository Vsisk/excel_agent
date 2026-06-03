import numpy

for name, fallback in {
    "short": int,
    "ushort": int,
    "intc": int,
    "uintc": int,
    "int_": int,
    "uint": int,
    "longlong": int,
    "ulonglong": int,
    "half": float,
    "float16": float,
    "single": float,
    "double": float,
    "longdouble": float,
    "int8": int,
    "int16": int,
    "int32": int,
    "int64": int,
    "uint8": int,
    "uint16": int,
    "uint32": int,
    "uint64": int,
    "intp": int,
    "uintp": int,
    "float32": float,
    "float64": float,
    "bool_": bool,
    "floating": float,
    "integer": int,
}.items():
    if not hasattr(numpy, name):
        setattr(numpy, name, fallback)

from openpyxl import Workbook

from agent.excel_agent.table_markdown_extractor import get_table_md_by_bbox, get_table_md_by_cell_range


def test_get_table_md_by_cell_range_reads_zero_based_half_open_range(tmp_path):
    path = tmp_path / "cell_range.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Invoice"
    ws["B1"] = "Amount"
    ws["A2"] = "INV-001"
    ws["B2"] = 120
    ws["A5"] = "Item"
    ws["B5"] = "Qty"
    wb.save(path)

    table_md = get_table_md_by_cell_range(
        str(path),
        sheet_number=1,
        cell_range={"start_row": 0, "end_row": 2, "start_col": 0, "end_col": 2},
    )

    assert table_md == "| Invoice | Amount |\n| INV-001 | 120 |"


def test_get_table_md_by_cell_range_accepts_output_bbox_keys(tmp_path):
    path = tmp_path / "bbox_keys.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "SingleRow"
    ws["A1"] = "Item"
    ws["B1"] = "Qty"
    ws["C1"] = "Total"
    wb.save(path)

    table_md = get_table_md_by_cell_range(
        str(path),
        sheet_number=1,
        cell_range={"left": 0, "right": 3, "top": 0, "bottom": 1},
    )

    assert table_md == "- Item Qty Total"


def test_get_table_md_by_bbox_returns_exact_one_based_range(tmp_path):
    path = tmp_path / "bbox_exact_range.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Tables"
    ws["A1"] = "Invoice"
    ws["B1"] = "Amount"
    ws["A2"] = "INV-001"
    ws["B2"] = 120
    ws["A5"] = "Item"
    ws["B5"] = "Qty"
    ws["A6"] = "Storage"
    ws["B6"] = 3
    wb.save(path)

    table_md = get_table_md_by_bbox(str(path), sheet_number=1, bbox=[1, 5, 1, 6])

    assert table_md == "- Item\n- Storage"


def test_get_table_md_by_bbox_preserves_requested_cross_table_range(tmp_path):
    path = tmp_path / "bbox_cross_table_range.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Tables"
    ws["A1"] = "Invoice"
    ws["B1"] = "Amount"
    ws["A2"] = "INV-001"
    ws["B2"] = 120
    ws["E1"] = "Item"
    ws["F1"] = "Qty"
    ws["E2"] = "Storage"
    ws["F2"] = 3
    wb.save(path)

    table_md = get_table_md_by_bbox(str(path), sheet_number=1, bbox=[1, 1, 6, 2])

    assert table_md == "\n".join(
        [
            "| Invoice | Amount |  |  | Item | Qty |",
            "| INV-001 | 120 |  |  | Storage | 3 |",
        ]
    )
