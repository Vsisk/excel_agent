from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import xlrd

from .models import CellRange, ExcelWorkbookDict, SheetInfo, SheetInfoDict


class ExcelReader:
    """Read workbook metadata and stream sheet rows with openpyxl."""

    def __init__(self, file_path: str):
        self.file_path = str(Path(file_path))
        self.file_suffix = Path(file_path).suffix.lower()
        self._workbook = None
        self._sheet_info_by_name: dict[str, SheetInfo] = {}
        self._sheet_info_by_id: dict[str, SheetInfo] = {}

    def read(self) -> ExcelWorkbookDict:
        workbook = self._ensure_open()
        sheet_list: list[SheetInfoDict] = []
        self._sheet_info_by_name.clear()
        self._sheet_info_by_id.clear()

        if self._is_xls:
            for index, sheet_name in enumerate(workbook.sheet_names()):
                sheet = workbook.sheet_by_name(sheet_name)
                info = SheetInfo(
                    sheet_id=f"sheet_{index + 1}",
                    sheet_name=sheet_name,
                    sheet_index=index,
                    max_row=sheet.nrows,
                    max_col=sheet.ncols,
                    merged_cells=[],
                )
                self._sheet_info_by_name[sheet_name] = info
                self._sheet_info_by_id[info.sheet_id] = info
                sheet_list.append(info.to_dict())
        else:
            for index, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]
                info = SheetInfo(
                    sheet_id=f"sheet_{index + 1}",
                    sheet_name=sheet_name,
                    sheet_index=index,
                    max_row=sheet.max_row or 0,
                    max_col=sheet.max_column or 0,
                    merged_cells=self._merged_ranges(sheet),
                )
                self._sheet_info_by_name[sheet_name] = info
                self._sheet_info_by_id[info.sheet_id] = info
                sheet_list.append(info.to_dict())

        return {"sheet_list": sheet_list}

    def get_sheet_info(self, sheet_name: str) -> SheetInfo:
        if not self._sheet_info_by_name:
            self.read()
        return self._sheet_info_by_name[sheet_name]

    def iter_rows(
        self,
        sheet_name: str,
        *,
        min_row: int = 1,
        max_row: int | None = None,
        min_col: int = 1,
        max_col: int | None = None,
    ) -> Iterator[list[Any]]:
        workbook = self._ensure_open()
        if self._is_xls:
            sheet = workbook.sheet_by_name(sheet_name)
            row_end = max_row if max_row is not None else sheet.nrows
            col_end = max_col if max_col is not None else sheet.ncols
            for row_index in range(min_row, row_end + 1):
                values = []
                for col_index in range(min_col, col_end + 1):
                    if row_index <= sheet.nrows and col_index <= sheet.ncols:
                        values.append(self._normalize_xls_value(sheet.cell_value(row_index - 1, col_index - 1)))
                    else:
                        values.append(None)
                yield values
            return

        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            yield list(row)

    def read_range(self, sheet_name: str, cell_range: CellRange) -> list[list[Any]]:
        return list(
            self.iter_rows(
                sheet_name,
                min_row=cell_range.start_row,
                max_row=cell_range.end_row,
                min_col=cell_range.start_col,
                max_col=cell_range.end_col,
            )
        )

    def close(self) -> None:
        if self._workbook is not None:
            close = getattr(self._workbook, "close", None)
            if close is not None:
                close()
            self._workbook = None

    def _ensure_open(self):
        if self._workbook is None:
            if self._is_xls:
                self._workbook = xlrd.open_workbook(self.file_path, on_demand=True)
            else:
                self._workbook = load_workbook(self.file_path, read_only=True, data_only=True)
        return self._workbook

    @property
    def _is_xls(self) -> bool:
        return self.file_suffix == ".xls"

    @staticmethod
    def _normalize_xls_value(value: Any) -> Any:
        if value == "":
            return None
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    @staticmethod
    def _merged_ranges(sheet: Any) -> list[str]:
        merged = getattr(sheet, "merged_cells", None)
        ranges = getattr(merged, "ranges", []) if merged is not None else []
        return [str(cell_range) for cell_range in ranges]
